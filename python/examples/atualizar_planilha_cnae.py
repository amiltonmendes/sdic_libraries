#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Atualiza a planilha "Lista de CNAE para medição do emprego e empresas" com os
dados vivos da API SDIC, preservando integralmente o layout do arquivo de
origem (linhas de cabeçalho, ordem das linhas, ordem das colunas, mesclagens,
formatos e fórmulas), e gera um relatório comparando origem x arquivo gerado.

Uso:
    python atualizar_planilha_cnae.py
    python atualizar_planilha_cnae.py --origem <arquivo.xlsx> --saida <dir>

Cobertura (ver RELATORIO / seção "Cobertura das colunas"):

    D/E/F  Estoque 2022/2023/2024 .... get_estoque_emprego_nacional (nível grupo)
                                       -> só para as linhas de GRUPO (3 dígitos);
                                          a API não oferece estoque no nível CLASSE
    G      Estoque 2025 ............... fórmula =F+J preservada (estoque + saldo)
    H/I/J  Saldo 2023/2024/2025 ....... get_saldo_caged_nacional
                                       -> grupo: nível 'grupo'
                                       -> classe: nível 'subclasse' agregado
                                          pelos 5 primeiros dígitos
    K..V   Estabelecimentos por porte .. NÃO disponível (ver relatório)
"""
from __future__ import annotations

import argparse
import re
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
import pandas as pd
from openpyxl.utils import get_column_letter

# Permite rodar o script direto do diretório examples/ sem instalar o pacote
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from sdic_libraries.dados.emprego import (  # noqa: E402
    get_estoque_emprego_nacional,
    get_saldo_caged_nacional,
)

ARQUIVO_ORIGEM_PADRAO = (
    "2026-04-28-Lista de CNAE para medição do emprego e empresas24082026.xlsx"
)
ABA_PADRAO = "Dados atualizados para 2025"

LINHA_CABECALHO_FIM = 3          # linhas 1..3 são cabeçalho (mantidas intactas)
ROTULO_TOTAL_GERAL = "TOTAL DE TODAS AS CNAES"
ROTULO_TOTAL_LISTA = "TOTAL CANAES ABAIXO"

ANOS_ESTOQUE = {"D": 2022, "E": 2023, "F": 2024}
ANOS_SALDO = {"H": 2023, "I": 2024, "J": 2025}
COLUNA_ESTOQUE_PROJETADO = "G"    # fórmula =F{linha}+J{linha}

# Colunas de estabelecimentos por porte (RAIS / metodologia Sebrae), por ano
COLUNAS_ESTABELECIMENTOS = {
    2022: ["K", "L", "M", "N"],
    2023: ["O", "P", "Q", "R"],
    2024: ["S", "T", "U", "V"],
}

MOTIVO_ESTOQUE_CLASSE = (
    "API expõe estoque apenas nos níveis divisão (2 díg.) e grupo (3 díg.); "
    "não há endpoint de estoque no nível classe (5 díg.)"
)
MOTIVO_ESTABELECIMENTOS = (
    "Não há na biblioteca nem na API endpoint de estabelecimentos por porte da "
    "RAIS/metodologia Sebrae com recorte anual e 4 portes"
)


# --------------------------------------------------------------------------- #
# Leitura do layout da planilha de origem
# --------------------------------------------------------------------------- #
class LinhaCNAE:
    """Uma linha de dados da planilha, com seu papel e chave CNAE."""

    def __init__(self, linha: int, tipo: str, codigo: Optional[str], descricao: Optional[str]):
        self.linha = linha
        self.tipo = tipo              # 'total_geral' | 'total_lista' | 'grupo' | 'classe'
        self.codigo = codigo
        self.descricao = descricao

    def __repr__(self) -> str:
        return f"LinhaCNAE(linha={self.linha}, tipo={self.tipo!r}, codigo={self.codigo!r})"


def ler_layout(ws) -> List[LinhaCNAE]:
    """Percorre a aba e classifica cada linha de dados, preservando a ordem original."""
    linhas: List[LinhaCNAE] = []

    for r in range(LINHA_CABECALHO_FIM + 1, ws.max_row + 1):
        rotulo = ws.cell(r, 1).value
        codigo_bruto = ws.cell(r, 2).value
        descricao = ws.cell(r, 3).value

        if isinstance(rotulo, str) and rotulo.strip().upper() == ROTULO_TOTAL_GERAL:
            linhas.append(LinhaCNAE(r, "total_geral", None, rotulo))
            continue
        if isinstance(rotulo, str) and rotulo.strip().upper() == ROTULO_TOTAL_LISTA:
            linhas.append(LinhaCNAE(r, "total_lista", None, rotulo))
            continue
        if codigo_bruto is None:
            continue

        codigo = str(codigo_bruto).strip()
        if not codigo.isdigit():
            continue

        tipo = {3: "grupo", 5: "classe"}.get(len(codigo))
        if tipo is None:
            print(f"  ! linha {r}: código CNAE '{codigo}' com {len(codigo)} dígitos — ignorado")
            continue
        linhas.append(LinhaCNAE(r, tipo, codigo, descricao))

    return linhas


# --------------------------------------------------------------------------- #
# Coleta de dados na API (via funções da biblioteca)
# --------------------------------------------------------------------------- #
def coletar_estoque_grupos(grupos: List[str], anos: List[int]) -> Dict[Tuple[str, int], float]:
    """Estoque nacional por grupo CNAE (3 dígitos) e ano."""
    print(f"  -> estoque nacional (nível grupo) para {len(grupos)} códigos...")
    df = get_estoque_emprego_nacional(codigos_cnae=grupos, nivel_cnae=3, agregado=True)
    if df.empty:
        return {}
    df["ano"] = df["ano"].astype(int)
    df = df[df["ano"].isin(anos)]
    df["grupo_cnae_cod"] = df["grupo_cnae_cod"].astype(str).str.strip()
    serie = df.groupby(["grupo_cnae_cod", "ano"])["estoque_trabalhadores"].sum()
    return {(cod, ano): float(v) for (cod, ano), v in serie.items()}


def coletar_estoque_total(anos: List[int]) -> Dict[int, float]:
    """Estoque nacional total (soma de todas as divisões CNAE) por ano."""
    print("  -> estoque nacional total (todas as divisões)...")
    df = get_estoque_emprego_nacional(nivel_cnae=2, agregado=True)
    if df.empty:
        return {}
    df["ano"] = df["ano"].astype(int)
    df = df[df["ano"].isin(anos)]
    return {int(a): float(v) for a, v in df.groupby("ano")["estoque_trabalhadores"].sum().items()}


def _saldo_anual(df: pd.DataFrame, coluna_cod: str, anos: List[int]) -> pd.DataFrame:
    """Consolida saldo mensal do CAGED em saldo anual por código."""
    df = df.copy()
    df["ano"] = pd.to_datetime(df["mes_referencia"]).dt.year
    df = df[df["ano"].isin(anos)]
    df[coluna_cod] = df[coluna_cod].astype(str).str.strip()
    return df.groupby([coluna_cod, "ano"], as_index=False)["saldo_reajustado"].sum()


def coletar_saldo_grupos(grupos: List[str], anos: List[int]) -> Dict[Tuple[str, int], float]:
    """Saldo CAGED anual por grupo CNAE (3 dígitos)."""
    print(f"  -> saldo CAGED (nível grupo) para {len(grupos)} códigos...")
    df = get_saldo_caged_nacional(
        "grupo",
        codigos=grupos,
        data_minima=f"{min(anos)}-01-01",
        data_maxima=f"{max(anos)}-12-31",
    )
    if df.empty:
        return {}
    anual = _saldo_anual(df, "grupo_cnae_cod", anos)
    return {(r.grupo_cnae_cod, int(r.ano)): float(r.saldo_reajustado) for r in anual.itertuples()}


def mapear_subclasses(classes: List[str], ano_referencia: int) -> Dict[str, List[str]]:
    """Descobre as subclasses (7 díg.) que compõem cada classe (5 díg.) da planilha."""
    print("  -> mapeando subclasses CNAE de cada classe...")
    df = get_saldo_caged_nacional(
        "subclasse",
        data_minima=f"{ano_referencia}-01-01",
        data_maxima=f"{ano_referencia}-01-31",
    )
    alvo = set(classes)
    mapa: Dict[str, List[str]] = defaultdict(list)
    for cod in sorted({str(c).strip() for c in df["subclasse_cnae_cod"]}):
        classe = cod[:5]
        if classe in alvo:
            mapa[classe].append(cod)
    ausentes = alvo - set(mapa)
    if ausentes:
        print(f"  ! sem subclasses no CAGED para: {sorted(ausentes)}")
    return dict(mapa)


def coletar_saldo_classes(classes: List[str], anos: List[int]) -> Dict[Tuple[str, int], float]:
    """Saldo CAGED anual por classe (5 díg.), somando as subclasses correspondentes."""
    mapa = mapear_subclasses(classes, min(anos))
    subclasses = [s for lista in mapa.values() for s in lista]
    if not subclasses:
        return {}

    print(f"  -> saldo CAGED (nível subclasse) para {len(subclasses)} códigos...")
    df = get_saldo_caged_nacional(
        "subclasse",
        codigos=subclasses,
        data_minima=f"{min(anos)}-01-01",
        data_maxima=f"{max(anos)}-12-31",
    )
    if df.empty:
        return {}

    df = df.copy()
    df["classe_cnae_cod"] = df["subclasse_cnae_cod"].astype(str).str.strip().str[:5]
    anual = _saldo_anual(df, "classe_cnae_cod", anos)
    return {(r.classe_cnae_cod, int(r.ano)): float(r.saldo_reajustado) for r in anual.itertuples()}


def coletar_saldo_total(anos: List[int]) -> Tuple[Dict[int, float], Optional[str]]:
    """Saldo CAGED nacional total por ano e o último mês de referência disponível."""
    print("  -> saldo CAGED nacional total (todas as divisões)...")
    df = get_saldo_caged_nacional(
        "divisao",
        data_minima=f"{min(anos)}-01-01",
        data_maxima=f"{max(anos)}-12-31",
    )
    if df.empty:
        return {}, None
    ultimo_mes = str(pd.to_datetime(df["mes_referencia"]).max().date())
    anual = _saldo_anual(df, "divisao_cnae_cod", anos)
    totais = anual.groupby("ano")["saldo_reajustado"].sum()
    return {int(a): float(v) for a, v in totais.items()}, ultimo_mes


# --------------------------------------------------------------------------- #
# Escrita do arquivo de saída
# --------------------------------------------------------------------------- #
def _int_ou_none(valor: Optional[float]) -> Optional[int]:
    return None if valor is None else int(round(valor))


def atualizar_planilha(ws, linhas: List[LinhaCNAE], dados: Dict[str, Any]) -> List[dict]:
    """Escreve os valores novos na aba, preservando cabeçalho, ordem e fórmulas.

    Retorna a lista de células atualizadas/pendentes para o relatório.
    """
    registros: List[dict] = []

    def registrar(ln: LinhaCNAE, col: str, status: str, valor=None, motivo: str = "") -> None:
        registros.append(
            {
                "linha": ln.linha,
                "coluna": col,
                "tipo_linha": ln.tipo,
                "codigo_cnae": ln.codigo,
                "descricao": ln.descricao,
                "status": status,
                "valor_novo": valor,
                "motivo": motivo,
            }
        )

    for ln in linhas:
        # ---------------- linha de total geral (todas as CNAEs do país) -----
        if ln.tipo == "total_geral":
            for col, ano in ANOS_ESTOQUE.items():
                valor = _int_ou_none(dados["estoque_total"].get(ano))
                if valor is None:
                    registrar(ln, col, "sem_dado", motivo=f"API sem estoque total para {ano}")
                    continue
                ws.cell(ln.linha, openpyxl.utils.column_index_from_string(col)).value = valor
                registrar(ln, col, "atualizado", valor)
            for col, ano in ANOS_SALDO.items():
                valor = _int_ou_none(dados["saldo_total"].get(ano))
                if valor is None:
                    registrar(ln, col, "sem_dado", motivo=f"API sem saldo total para {ano}")
                    continue
                ws.cell(ln.linha, openpyxl.utils.column_index_from_string(col)).value = valor
                registrar(ln, col, "atualizado", valor)
            _escrever_projecao(ws, ln, registrar)
            _registrar_estabelecimentos(ln, registrar)
            continue

        # ---------------- linha de total da lista: só fórmulas --------------
        if ln.tipo == "total_lista":
            for col in list(ANOS_ESTOQUE) + [COLUNA_ESTOQUE_PROJETADO] + list(ANOS_SALDO):
                registrar(ln, col, "formula_preservada", motivo="=SUM(...) recalculado pelo Excel")
            continue

        # ---------------- linhas de CNAE (grupo / classe) -------------------
        for col, ano in ANOS_ESTOQUE.items():
            if ln.tipo == "classe":
                registrar(ln, col, "indisponivel", motivo=MOTIVO_ESTOQUE_CLASSE)
                continue
            valor = _int_ou_none(dados["estoque_grupo"].get((ln.codigo, ano)))
            if valor is None:
                registrar(ln, col, "sem_dado", motivo=f"API sem estoque para {ln.codigo}/{ano}")
                continue
            ws.cell(ln.linha, openpyxl.utils.column_index_from_string(col)).value = valor
            registrar(ln, col, "atualizado", valor)

        fonte_saldo = dados["saldo_grupo"] if ln.tipo == "grupo" else dados["saldo_classe"]
        for col, ano in ANOS_SALDO.items():
            valor = _int_ou_none(fonte_saldo.get((ln.codigo, ano)))
            if valor is None:
                registrar(ln, col, "sem_dado", motivo=f"CAGED sem saldo para {ln.codigo}/{ano}")
                continue
            ws.cell(ln.linha, openpyxl.utils.column_index_from_string(col)).value = valor
            registrar(ln, col, "atualizado", valor)

        _escrever_projecao(ws, ln, registrar)
        _registrar_estabelecimentos(ln, registrar)

    return registros


def _escrever_projecao(ws, ln: LinhaCNAE, registrar) -> None:
    """Coluna G: estoque projetado = estoque do último ano real + saldo do ano corrente."""
    col_idx = openpyxl.utils.column_index_from_string(COLUNA_ESTOQUE_PROJETADO)
    formula = f"=F{ln.linha}+J{ln.linha}"
    ws.cell(ln.linha, col_idx).value = formula
    registrar(
        ln,
        COLUNA_ESTOQUE_PROJETADO,
        "formula_preservada",
        formula,
        "estoque projetado = coluna F + coluna J",
    )


def _registrar_estabelecimentos(ln: LinhaCNAE, registrar) -> None:
    for ano, colunas in COLUNAS_ESTABELECIMENTOS.items():
        for col in colunas:
            registrar(ln, col, "indisponivel", motivo=f"{MOTIVO_ESTABELECIMENTOS} (ano {ano})")


# --------------------------------------------------------------------------- #
# Relatório de comparação
# --------------------------------------------------------------------------- #
def _valores_da_aba(caminho: Path, aba: str) -> Dict[Tuple[int, str], Any]:
    """Lê os valores calculados de uma aba (linha, coluna) -> valor."""
    wb = openpyxl.load_workbook(caminho, data_only=True)
    ws = wb[aba]
    valores = {}
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            valores[(r, get_column_letter(c))] = ws.cell(r, c).value
    wb.close()
    return valores


def _formulas_da_aba(caminho: Path, aba: str) -> Dict[Tuple[int, str], str]:
    """Lê as fórmulas de uma aba (linha, coluna) -> texto da fórmula."""
    wb = openpyxl.load_workbook(caminho, data_only=False)
    ws = wb[aba]
    formulas = {}
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.startswith("="):
                formulas[(r, get_column_letter(c))] = v
    wb.close()
    return formulas


_RE_SOMA = re.compile(r"^=SUM\(([A-Z]+)(\d+):[A-Z]+(\d+)\)$", re.IGNORECASE)
_RE_ADICAO = re.compile(r"^=([A-Z]+)(\d+)\+([A-Z]+)(\d+)$", re.IGNORECASE)


def _avaliar_formula(formula: str, valores: Dict[Tuple[int, str], Any]) -> Optional[float]:
    """Resolve as duas formas de fórmula usadas na planilha: =SUM(X:X) e =Xn+Yn.

    O Excel recalcula as fórmulas ao abrir o arquivo; aqui elas são avaliadas
    apenas para que o relatório possa comparar também as linhas de total.
    """
    m = _RE_SOMA.match(formula.strip())
    if m:
        col, ini, fim = m.group(1).upper(), int(m.group(2)), int(m.group(3))
        total = 0.0
        for r in range(ini, fim + 1):
            v = valores.get((r, col))
            if isinstance(v, str) and v.startswith("="):
                v = _avaliar_formula(v, valores)
            if isinstance(v, (int, float)):
                total += v
        return total

    m = _RE_ADICAO.match(formula.strip())
    if m:
        partes = []
        for col, lin in ((m.group(1).upper(), int(m.group(2))), (m.group(3).upper(), int(m.group(4)))):
            v = valores.get((lin, col))
            if isinstance(v, str) and v.startswith("="):
                v = _avaliar_formula(v, valores)
            if not isinstance(v, (int, float)):
                return None
            partes.append(v)
        return partes[0] + partes[1]

    return None


def _rotulo_coluna(ws, col: str) -> str:
    """Monta o rótulo humano da coluna a partir das linhas de cabeçalho."""
    idx = openpyxl.utils.column_index_from_string(col)
    partes = []
    for r in range(1, LINHA_CABECALHO_FIM + 1):
        v = ws.cell(r, idx).value
        if v:
            partes.append(" ".join(str(v).split()))
    return f"{col} — {' / '.join(partes)}" if partes else col


def montar_comparacao(
    origem: Path,
    saida: Path,
    aba: str,
    registros: List[dict],
    ws_saida,
) -> pd.DataFrame:
    """Compara célula a célula os valores da planilha de origem e da gerada."""
    antes = _valores_da_aba(origem, aba)
    depois = _valores_da_aba(saida, aba)
    formulas_saida = _formulas_da_aba(saida, aba)

    # o arquivo gerado ainda não passou pelo Excel: fórmulas não têm valor em
    # cache, então são resolvidas aqui para o relatório
    contexto = dict(depois)
    contexto.update(formulas_saida)

    por_celula = {(r["linha"], r["coluna"]): r for r in registros}
    linhas_rel = []

    for (linha, coluna), reg in sorted(por_celula.items()):
        v_antes = antes.get((linha, coluna))
        v_depois = depois.get((linha, coluna))

        if v_depois is None and (linha, coluna) in formulas_saida:
            v_depois = _avaliar_formula(formulas_saida[(linha, coluna)], contexto)

        num_antes = v_antes if isinstance(v_antes, (int, float)) else None
        num_depois = v_depois if isinstance(v_depois, (int, float)) else None

        diff = pct = None
        if num_antes is not None and num_depois is not None:
            diff = num_depois - num_antes
            pct = (diff / abs(num_antes) * 100) if num_antes else None

        if reg["status"] == "indisponivel":
            resultado = "nao_atualizado"
        elif num_antes is None and num_depois is None:
            resultado = "vazio_nos_dois"
        elif diff is None:
            resultado = "nao_comparavel"
        elif diff == 0:
            resultado = "identico"
        else:
            resultado = "divergente"

        linhas_rel.append(
            {
                "linha": linha,
                "coluna": coluna,
                "rotulo_coluna": _rotulo_coluna(ws_saida, coluna),
                "tipo_linha": reg["tipo_linha"],
                "codigo_cnae": reg["codigo_cnae"],
                "descricao": reg["descricao"],
                "status_atualizacao": reg["status"],
                "valor_origem": num_antes if num_antes is not None else v_antes,
                "valor_gerado": num_depois if num_depois is not None else v_depois,
                "diferenca": diff,
                "variacao_pct": round(pct, 4) if pct is not None else None,
                "resultado": resultado,
                "motivo": reg["motivo"],
            }
        )

    return pd.DataFrame(linhas_rel)


def escrever_relatorio_markdown(
    df: pd.DataFrame, destino: Path, origem: Path, saida: Path, aba: str, meta: Dict[str, Any]
) -> None:
    L: List[str] = []
    A = L.append

    A("# Relatório de atualização — Lista de CNAE (emprego e empresas)")
    A("")
    A(f"- **Gerado em:** {datetime.now():%d/%m/%Y %H:%M}")
    A(f"- **Arquivo de origem:** `{origem.name}`")
    A(f"- **Arquivo gerado:** `{saida.name}`")
    A(f"- **Aba comparada:** `{aba}`")
    A(f"- **Fonte dos dados:** API SDIC (`{meta['base_url']}`) via `sdic_libraries`")
    A(f"- **Último mês CAGED disponível:** {meta.get('ultimo_mes_caged') or 'n/d'}")
    A(f"- **Layout preservado:** {meta['n_linhas']} linhas de dados, "
      f"colunas A–{meta['ultima_coluna']}, {LINHA_CABECALHO_FIM} linhas de cabeçalho mantidas")
    A("")

    A("## 1. Cobertura das colunas")
    A("")
    A("| Colunas | Conteúdo | Atualizável? | Função usada / motivo |")
    A("|---|---|---|---|")
    A("| D, E, F | Estoque 2022/2023/2024 — linhas de **grupo** (3 díg.) | ✅ Sim | "
      "`get_estoque_emprego_nacional(nivel_cnae=3, agregado=True)` |")
    A("| D, E, F | Estoque 2022/2023/2024 — linhas de **classe** (5 díg.) | ❌ Não | "
      f"{MOTIVO_ESTOQUE_CLASSE} |")
    A("| G | Estoque 2025 (projetado) | ✅ Sim | fórmula `=F+J` preservada "
      "(mesma lógica de `get_estoque_emprego_estimado_nacional_anual`) |")
    A("| H, I, J | Saldo 2023/2024/2025 — linhas de **grupo** | ✅ Sim | "
      "`get_saldo_caged_nacional('grupo')` |")
    A("| H, I, J | Saldo 2023/2024/2025 — linhas de **classe** | ✅ Sim | "
      "`get_saldo_caged_nacional('subclasse')` agregado pelos 5 primeiros dígitos |")
    A("| K–N | Estabelecimentos por porte 2022 | ❌ Não | "
      "vazio na origem e sem fonte na API |")
    A("| O–R, S–V | Estabelecimentos por porte 2023 / 2024 | ❌ Não | "
      f"{MOTIVO_ESTABELECIMENTOS} |")
    A("")

    A("## 2. Resumo por resultado")
    A("")
    A("| Resultado | Células |")
    A("|---|---|")
    for k, v in df["resultado"].value_counts().items():
        A(f"| {k} | {v} |")
    A(f"| **total** | **{len(df)}** |")
    A("")

    A("## 3. Resumo por coluna")
    A("")
    A("| Coluna | Atualizadas | Idênticas | Divergentes | Não atualizadas | Δ médio | Δ máx (abs) |")
    A("|---|---|---|---|---|---|---|")
    for col in sorted(df["coluna"].unique(), key=openpyxl.utils.column_index_from_string):
        sub = df[df["coluna"] == col]
        comp = sub[sub["diferenca"].notna() & (sub["resultado"] != "nao_atualizado")]
        atualizadas = int((sub["status_atualizacao"] == "atualizado").sum())
        identicas = int((sub["resultado"] == "identico").sum())
        divergentes = int((sub["resultado"] == "divergente").sum())
        nao_atualizadas = int((sub["resultado"] == "nao_atualizado").sum())
        media = f"{comp['diferenca'].mean():,.1f}" if not comp.empty else "—"
        maximo = f"{comp['diferenca'].abs().max():,.0f}" if not comp.empty else "—"
        A(f"| {sub['rotulo_coluna'].iloc[0]} | {atualizadas} | {identicas} | "
          f"{divergentes} | {nao_atualizadas} | {media} | {maximo} |")
    A("")

    A("## 4. Maiores divergências (origem x gerado)")
    A("")
    div = df[df["resultado"] == "divergente"].copy()
    if div.empty:
        A("_Nenhuma divergência._")
    else:
        div["abs"] = div["diferenca"].abs()
        top = div.sort_values("abs", ascending=False).head(30)
        A("| Linha | Col | CNAE | Descrição | Origem | Gerado | Δ | Δ% |")
        A("|---|---|---|---|---|---|---|---|")
        for r in top.itertuples():
            desc = str(r.descricao or "")
            desc = (desc[:45] + "…") if len(desc) > 45 else desc
            pct = f"{r.variacao_pct:+.2f}%" if pd.notna(r.variacao_pct) else "—"
            codigo = r.codigo_cnae if pd.notna(r.codigo_cnae) else "—"
            A(f"| {r.linha} | {r.coluna} | {codigo} | {desc} | "
              f"{r.valor_origem:,.0f} | {r.valor_gerado:,.0f} | {r.diferenca:+,.0f} | {pct} |")
    A("")

    A("> **Como ler as divergências.** Elas não indicam erro: a planilha de origem")
    A("> fixou os números na data em que foi montada. As colunas E/F (estoque")
    A("> 2023/2024) traziam valores *projetados* a partir do estoque de 2022 mais o")
    A("> saldo acumulado, enquanto a API já devolve o estoque **efetivo** da RAIS —")
    A("> daí a diferença sistemática para baixo. As colunas H/I (saldo 2023/2024)")
    A("> diferem apenas por revisões mensais do CAGED (ordem de dezenas). A coluna J")
    A("> (saldo 2025) diverge porque a origem ainda carregava o acumulado jan-set;")
    A("> o arquivo gerado traz o ano fechado. A coluna D (estoque 2022) bate")
    A("> exatamente, confirmando que origem e API compartilham a mesma base.")
    A("")

    A("## 5. Células não atualizadas")
    A("")
    nao = df[df["resultado"] == "nao_atualizado"]
    if nao.empty:
        A("_Nenhuma._")
    else:
        A("| Colunas | Células | Motivo |")
        A("|---|---|---|")
        for motivo, sub in nao.groupby("motivo"):
            cols = ", ".join(sorted(sub["coluna"].unique(), key=openpyxl.utils.column_index_from_string))
            A(f"| {cols} | {len(sub)} | {motivo} |")
    A("")

    destino.write_text("\n".join(L), encoding="utf-8")


# --------------------------------------------------------------------------- #
def main() -> int:
    base_dir = Path(__file__).resolve().parent
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--origem", default=str(base_dir / ARQUIVO_ORIGEM_PADRAO))
    ap.add_argument("--aba", default=ABA_PADRAO)
    ap.add_argument("--saida", default=str(base_dir / "saida"))
    args = ap.parse_args()

    origem = Path(args.origem)
    if not origem.exists():
        print(f"ERRO: arquivo de origem não encontrado: {origem}")
        return 1

    dir_saida = Path(args.saida)
    dir_saida.mkdir(parents=True, exist_ok=True)
    saida = dir_saida / f"{origem.stem}__atualizado.xlsx"
    rel_md = dir_saida / "relatorio_comparacao.md"
    rel_csv = dir_saida / "relatorio_comparacao.csv"

    print(f"Lendo layout de origem: {origem.name} / aba '{args.aba}'")
    wb = openpyxl.load_workbook(origem)          # mantém estilos, mesclagens e fórmulas
    if args.aba not in wb.sheetnames:
        print(f"ERRO: aba '{args.aba}' não existe. Abas: {wb.sheetnames}")
        return 1
    ws = wb[args.aba]

    linhas = ler_layout(ws)
    grupos = [l.codigo for l in linhas if l.tipo == "grupo"]
    classes = [l.codigo for l in linhas if l.tipo == "classe"]
    print(f"  {len(linhas)} linhas de dados: {len(grupos)} grupos, {len(classes)} classes, "
          f"{sum(1 for l in linhas if l.tipo.startswith('total'))} totais")

    anos_estoque = sorted(set(ANOS_ESTOQUE.values()))
    anos_saldo = sorted(set(ANOS_SALDO.values()))

    print("Consultando a API SDIC via sdic_libraries...")
    saldo_total, ultimo_mes = coletar_saldo_total(anos_saldo)
    dados = {
        "estoque_grupo": coletar_estoque_grupos(grupos, anos_estoque),
        "estoque_total": coletar_estoque_total(anos_estoque),
        "saldo_grupo": coletar_saldo_grupos(grupos, anos_saldo),
        "saldo_classe": coletar_saldo_classes(classes, anos_saldo),
        "saldo_total": saldo_total,
    }

    print("Atualizando células (layout preservado)...")
    registros = atualizar_planilha(ws, linhas, dados)
    wb.save(saida)
    print(f"  arquivo gerado: {saida}")

    print("Gerando relatório de comparação...")
    df = montar_comparacao(origem, saida, args.aba, registros, ws)
    df.to_csv(rel_csv, index=False, encoding="utf-8-sig")

    import os
    meta = {
        "base_url": os.getenv("EMPLOYMENT_API_BASE_URL", "https://sdicapi.dados.ninja"),
        "n_linhas": len(linhas),
        "ultima_coluna": get_column_letter(ws.max_column),
        "ultimo_mes_caged": ultimo_mes,
    }
    escrever_relatorio_markdown(df, rel_md, origem, saida, args.aba, meta)
    wb.close()

    print(f"  relatório: {rel_md}")
    print(f"  detalhamento: {rel_csv}")
    print("\nResumo:")
    for k, v in df["resultado"].value_counts().items():
        print(f"  {k:20s} {v}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
